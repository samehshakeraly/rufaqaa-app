"""XLSX building for report exports.

Pure functions: title + rows in, bytes out. No DB access — data fetching and
tenant scoping belong to the caller.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ثقة — brand accent used for header rows, matching the PDF template.
_HEADER_BG = "FF769FCD"
_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 60

# Each sheet is (sheet_title, header_labels, rows).
Sheet = tuple[str, list[str], list[list[object]]]


def build_report_xlsx(title: str, sheets: list[Sheet]) -> bytes:
    """Build an RTL-oriented workbook and return it as XLSX bytes."""
    wb = Workbook()
    # Workbook() starts with one default sheet; drop it and create our own.
    default = wb.active
    if default is not None:
        wb.remove(default)
    wb.properties.title = title

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for sheet_title, headers, rows in sheets:
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.rightToLeft = True

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row in rows:
            ws.append(row)

        # Auto-ish widths: size each column to its longest rendered value,
        # clamped so one long cell can't blow up the layout.
        for col_idx, header in enumerate(headers, start=1):
            longest = len(header)
            for row in rows:
                if col_idx <= len(row):
                    longest = max(longest, len(str(row[col_idx - 1])))
            width = min(max(longest + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

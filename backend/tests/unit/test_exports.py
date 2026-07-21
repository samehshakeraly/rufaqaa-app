"""Arabic proof-of-concept for the shared export rendering layer (PDF + XLSX)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from weasyprint import HTML

from app.services.exports.excel import build_report_xlsx
from app.services.exports.pdf import PACKAGE_DIR, render_html, render_report_pdf

ARABIC_TITLE = "تقرير تجريبي"
ARABIC_HEADERS = ["الاسم", "العمر", "المبلغ"]
NUMERIC_ROWS = [
    ["أحمد محمد", 7, 150],
    ["فاطمة علي", 9, 200],
    ["يوسف خالد", 11, 175],
]


def _base_context() -> dict[str, object]:
    return {
        "title": ARABIC_TITLE,
        "generated_at": "2026-07-21 12:00",
        "headers": ARABIC_HEADERS,
        "rows": NUMERIC_ROWS,
    }


def test_render_html_contains_arabic_title_and_rtl() -> None:
    html = render_html("base.html", _base_context())
    assert ARABIC_TITLE in html
    assert 'dir="rtl"' in html
    assert 'lang="ar"' in html


def test_render_report_pdf_produces_valid_single_page_pdf() -> None:
    pdf = render_report_pdf("base.html", _base_context())
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 1000

    html = render_html("base.html", _base_context())
    document = HTML(string=html, base_url=str(PACKAGE_DIR)).render()
    assert len(document.pages) == 1


def test_build_report_xlsx_arabic_roundtrip() -> None:
    sheet_title = "الأيتام"
    arabic_value = "أحمد محمد"
    data = build_report_xlsx(
        ARABIC_TITLE,
        [(sheet_title, ARABIC_HEADERS, [[arabic_value, 7, 150]])],
    )
    assert data.startswith(b"PK")

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == [sheet_title]
    ws = wb[sheet_title]
    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(row=1, column=1).value == ARABIC_HEADERS[0]
    assert ws.cell(row=2, column=1).value == arabic_value
    assert ws.cell(row=2, column=2).value == 7

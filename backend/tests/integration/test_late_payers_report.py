"""GET /reports/late-payers — PR-R2 export endpoint.

Covers: valid PDF (%PDF) and XLSX (PK) bytes, the grace boundary on the
late derivation, tenant isolation (a second org's arrears never appear),
the finance/admin role gate (403 for donor / partner_staff /
marketing_manager), the summary row, and the privacy gate — no orphan PII
anywhere in the output rows.
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime, timedelta
from io import BytesIO

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import text

from app.core.constants import LATE_PAYMENT_GRACE_DAYS
from app.core.database import make_session
from app.core.security import hash_password
from app.models.organization import Organization
from app.models.user import User

URL = "/api/v1/reports/late-payers"


async def _seed_partner_id() -> str:
    async with make_session() as db:
        row = (
            await db.execute(
                text("SELECT id::text FROM partner_organizations WHERE code = 'DEV-PTN' LIMIT 1")
            )
        ).first()
        assert row is not None
        return row[0]


async def _make_donor(api: AsyncClient, h: dict[str, str], name: str = "كافل الاختبار") -> str:
    email = f"lp-{uuid.uuid4().hex[:8]}@example.com"
    r = await api.post("/api/v1/donors", json={"full_name": name, "email": email}, headers=h)
    assert r.status_code == 201, r.text
    return r.json()["id"]


async def _make_orphan(api: AsyncClient, h: dict[str, str]) -> dict:
    partner_id = await _seed_partner_id()
    unique = uuid.uuid4().hex[:6]
    r = await api.post(
        "/api/v1/orphans",
        json={
            "first_name": f"يتيم-سري-{unique}",
            "family_name": f"عائلة-سرية-{unique}",
            "date_of_birth": "2017-08-10",
            "gender": "M",
            "partner_organization_id": partner_id,
            "father_name": f"أب-سري-{unique}",
        },
        headers=h,
    )
    assert r.status_code == 201, r.text
    return r.json()


async def _make_sponsorship(
    api: AsyncClient,
    h: dict[str, str],
    donor_id: str,
    orphan_id: str,
    monthly: str = "20.00",
) -> dict:
    r = await api.post(
        "/api/v1/sponsorships",
        json={
            "donor_id": donor_id,
            "orphan_id": orphan_id,
            "monthly_amount": monthly,
            "currency": "KWD",
            "start_date": "2026-01-01",
        },
        headers=h,
    )
    assert r.status_code == 201, r.text
    return r.json()


async def _set_next_payment(
    sponsorship_id: str, days_past_due: int, status: str = "active"
) -> None:
    """Point next_payment_date `days_past_due` days into the past (negative
    values push it into the future)."""
    due = datetime.now(UTC).date() - timedelta(days=days_past_due)
    async with make_session() as db:
        await db.execute(
            text("UPDATE sponsorships SET next_payment_date = :d, status = :s WHERE id = :id"),
            {"d": due, "s": status, "id": sponsorship_id},
        )
        await db.commit()


def _xlsx_cells(content: bytes) -> list[str]:
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


async def _login_as(api: AsyncClient, admin_headers: dict[str, str], role: str) -> dict[str, str]:
    email = f"{role[:3]}-{uuid.uuid4().hex[:8]}@example.com"
    password = "longenoughpw1"
    r = await api.post(
        "/api/v1/users/invite",
        json={"email": email, "first_name": role.title(), "last_name": "User", "role": role},
        headers=admin_headers,
    )
    assert r.status_code == 201, r.text
    token = r.json()["invite_token"]
    r = await api.post("/api/v1/users/accept-invite", json={"token": token, "password": password})
    assert r.status_code == 200, r.text
    r = await api.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


async def _donor_headers(api: AsyncClient) -> dict[str, str]:
    email = f"dnr-{uuid.uuid4().hex[:8]}@example.com"
    r = await api.post(
        "/api/v1/auth/signup",
        json={"email": email, "password": "longenoughpw1", "full_name": "Donor User"},
    )
    assert r.status_code in (200, 201), r.text
    token = r.json()["debug_verify_token"]
    r = await api.post("/api/v1/auth/verify-email", json={"token": token})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


async def _new_org_token(api: AsyncClient, role: str = "finance") -> dict[str, str]:
    suffix = uuid.uuid4().hex[:8]
    email = f"{role[:3]}-{suffix}@other.example.com"
    password = "otherpw123456"
    async with make_session() as db:
        org = Organization(
            code=f"LPO-{suffix[:6].upper()}",
            name_ar="منظمة أخرى",
            name_en="Other Org",
            org_type="standalone",
            deployment_mode="self_hosted",
            country_code="KW",
        )
        db.add(org)
        await db.flush()
        db.add(
            User(
                organization_id=org.id,
                email=email,
                password_hash=hash_password(password),
                first_name="Other",
                last_name="Finance",
                role=role,
                status="active",
            )
        )
        await db.commit()
    r = await api.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


# ── Format smoke tests ──────────────────────────────────────────────────


async def test_pdf_export_returns_valid_pdf(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    donor_id = await _make_donor(api, auth_headers)
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"])
    await _set_next_payment(sp["id"], days_past_due=LATE_PAYMENT_GRACE_DAYS + 30)

    r = await api.get(f"{URL}?format=pdf", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/pdf")
    assert "attachment" in r.headers["content-disposition"]
    assert r.content.startswith(b"%PDF")


async def test_xlsx_export_returns_valid_workbook(
    api: AsyncClient, auth_headers: dict[str, str]
) -> None:
    donor_id = await _make_donor(api, auth_headers)
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"])
    await _set_next_payment(sp["id"], days_past_due=LATE_PAYMENT_GRACE_DAYS + 30)

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert r.content.startswith(b"PK")
    cells = _xlsx_cells(r.content)
    assert sp["code"] in cells


# ── Derivation: grace boundary + status gating ──────────────────────────


async def test_grace_boundary_and_current_sponsorships_excluded(
    api: AsyncClient, auth_headers: dict[str, str]
) -> None:
    donor_id = await _make_donor(api, auth_headers)
    o1 = await _make_orphan(api, auth_headers)
    o2 = await _make_orphan(api, auth_headers)
    o3 = await _make_orphan(api, auth_headers)

    late = await _make_sponsorship(api, auth_headers, donor_id, o1["id"])
    in_grace = await _make_sponsorship(api, auth_headers, donor_id, o2["id"])
    current = await _make_sponsorship(api, auth_headers, donor_id, o3["id"])

    # One day past the grace window → late.
    await _set_next_payment(late["id"], days_past_due=LATE_PAYMENT_GRACE_DAYS + 1)
    # Due exactly grace-days ago → still inside the grace window.
    await _set_next_payment(in_grace["id"], days_past_due=LATE_PAYMENT_GRACE_DAYS)
    # Not due yet.
    await _set_next_payment(current["id"], days_past_due=-30)

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    cells = _xlsx_cells(r.content)
    assert late["code"] in cells
    assert in_grace["code"] not in cells
    assert current["code"] not in cells


async def test_cancelled_and_paused_never_late(
    api: AsyncClient, auth_headers: dict[str, str]
) -> None:
    donor_id = await _make_donor(api, auth_headers)
    o1 = await _make_orphan(api, auth_headers)
    o2 = await _make_orphan(api, auth_headers)
    cancelled = await _make_sponsorship(api, auth_headers, donor_id, o1["id"])
    paused = await _make_sponsorship(api, auth_headers, donor_id, o2["id"])
    await _set_next_payment(cancelled["id"], days_past_due=120, status="cancelled")
    await _set_next_payment(paused["id"], days_past_due=120, status="paused")

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    cells = _xlsx_cells(r.content)
    assert cancelled["code"] not in cells
    assert paused["code"] not in cells


async def test_worker_flipped_overdue_status_included(
    api: AsyncClient, auth_headers: dict[str, str]
) -> None:
    donor_id = await _make_donor(api, auth_headers)
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"])
    await _set_next_payment(sp["id"], days_past_due=60, status="overdue")

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert sp["code"] in _xlsx_cells(r.content)


# ── Tenant isolation ────────────────────────────────────────────────────


async def test_other_orgs_late_sponsorships_never_appear(
    api: AsyncClient, auth_headers: dict[str, str]
) -> None:
    donor_id = await _make_donor(api, auth_headers)
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"])
    await _set_next_payment(sp["id"], days_past_due=90)

    other = await _new_org_token(api, role="finance")
    r = await api.get(f"{URL}?format=xlsx", headers=other)
    assert r.status_code == 200, r.text
    assert sp["code"] not in _xlsx_cells(r.content)


# ── Role gate ───────────────────────────────────────────────────────────


async def test_finance_role_allowed(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    finance = await _login_as(api, auth_headers, "finance")
    r = await api.get(f"{URL}?format=pdf", headers=finance)
    assert r.status_code == 200, r.text


async def test_non_finance_roles_forbidden(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    for role in ("partner_staff", "partner_manager", "marketing_manager"):
        headers = await _login_as(api, auth_headers, role)
        r = await api.get(URL, headers=headers)
        assert r.status_code == 403, f"{role}: {r.status_code} {r.text}"


async def test_donor_forbidden(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    donor = await _donor_headers(api)
    r = await api.get(URL, headers=donor)
    assert r.status_code == 403, r.text


async def test_anonymous_unauthorized(api: AsyncClient) -> None:
    r = await api.get(URL)
    assert r.status_code == 401, r.text


# ── Privacy: no orphan PII in the output ────────────────────────────────


async def test_no_orphan_pii_in_rows(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    donor_id = await _make_donor(api, auth_headers, name="كافل معروف")
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"])
    await _set_next_payment(sp["id"], days_past_due=45)

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    blob = " ".join(_xlsx_cells(r.content))

    assert sp["code"] in blob  # the row itself is present
    # The orphan must not be identifiable in any way: no name fragments, no
    # id, not even the orphan code (the report keys rows by sponsorship code).
    assert orphan["first_name"] not in blob
    assert orphan["family_name"] not in blob
    assert (orphan.get("father_name") or "أب-سري") not in blob
    assert orphan["id"] not in blob
    assert orphan["code"] not in blob


# ── Summary block ───────────────────────────────────────────────────────


async def test_summary_row_totals(api: AsyncClient, auth_headers: dict[str, str]) -> None:
    donor_id = await _make_donor(api, auth_headers)
    orphan = await _make_orphan(api, auth_headers)
    sp = await _make_sponsorship(api, auth_headers, donor_id, orphan["id"], monthly="30.00")
    # Due ~2 months ago → 3 installments behind (the due one + two elapsed).
    await _set_next_payment(sp["id"], days_past_due=65)

    r = await api.get(f"{URL}?format=xlsx", headers=auth_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    summary = rows[-1]
    assert "الإجمالي" in str(summary[0])
    # Our sponsorship contributes 90.00; other tests' rows may add more.
    assert float(summary[6]) >= 90.0

    sp_row = next(row for row in rows if row[0] == sp["code"])
    assert sp_row[5] == 3  # months behind
    assert float(sp_row[6]) == 90.0  # 3 × 30.00 in the org currency
